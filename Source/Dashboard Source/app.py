import sys
import os

# Add Shared and DataExploration
sys.path.append(os.path.join(os.getcwd(), 'Source', 'DataExploration'))
sys.path.append(os.path.join(os.getcwd(), 'Source', 'Shared'))

# Import các module cần thiết
from DataExploration_example import *
from DataExploration_01 import *
from DataExploration_02 import *
from DataExploration_03 import *
from Libraries import *
from Shared_Functions import *

#Dashboard Libraries
from flask import Flask, render_template, render_template, request, jsonify
import plotly.graph_objects as go
import json
import plotly
import requests
import openai
import io
from PIL import Image
import pytesseract  # OCR library
from PyPDF2 import PdfReader
from docx import Document
import openpyxl
import chardet

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

app = Flask(__name__)

# Set your OpenAI API key
openai.api_key = "Thay key vô rồi chạy"
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
def query_chatgpt(prompt):
    try:
        print(f"Sending to OpenAI: {prompt}")  # Debug
        response = openai.ChatCompletion.create(
            model="gpt-4o",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=2000,
            temperature=0.7,
        )
        print(f"Response from OpenAI: {response}")  # Debug
        return response['choices'][0]['message']['content']
    except Exception as e:
        print(f"Error querying OpenAI: {e}")  # Debug
        return f"Error querying OpenAI: {e}"

UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

#---------------------------------------------------- Không chỉnh sửa vùng ở trên. Trừ việc thêm thư viện vào.

@app.route('/')
def homepage():
    # Giả sử df là DataFrame chứa dataset
    num_rows = raw_df.shape[0]  # Số hàng
    num_columns = raw_df.shape[1]  # Số cột
    column_names = ', '.join(raw_df.columns)  # Tên các cột
    missing_values = raw_df.isnull().sum().to_dict()  # Thông tin về giá trị thiếu

    return render_template('index.html', 
                           num_rows=num_rows, 
                           num_columns=num_columns,
                           column_names=column_names,
                           missing_values=missing_values)

@app.route('/section1')
def section1():
    bar_avg_chart_json, pie_chart_json, histogram_charts, summary_stats,boxplot_chart_json = section_01_01()
    return render_template(
        'section1.html',
        bar_avg_chart=bar_avg_chart_json,
        pie_chart=pie_chart_json,
        histogram_charts=histogram_charts,
        summary_stats=summary_stats,
        boxplot_chart=boxplot_chart_json
    )

@app.route('/section2')
def section2():
    histogram_charts_json, summary_stats_json = section_02_01()
    grouped_bar_charts_json = section_02_02(df)
    density_plots_json = section_02_03(df)
    pie_charts_json = section_02_04(df)
    horizontal_bar_charts_json = section_02_05(df)
    student_group_donut_chart_json = section_02_06(df)
    return render_template(
        'section2.html',
        histogram_charts=histogram_charts_json,
        summary_stats=summary_stats_json,
        grouped_bar_charts=grouped_bar_charts_json,
        density_plots=density_plots_json,
        pie_charts=pie_charts_json,
        horizontal_bar_charts=horizontal_bar_charts_json,
        student_group_donut_chart=student_group_donut_chart_json
    )
                           
@app.route('/section3')
def section3():
    heatmap_json, strongest_correlations, weakest_correlations, launch_pad_stats, launch_pad_json, prediction_results = section_03_01()
    radar_chart_json = section_03_02()
    bar_chart_json = section_03_03()
    radar_chart_social_sciences_json = section_03_04()
    bar_chart_social_sciences_json = section_03_05()
    return render_template('section3.html', 
                        heatmap=heatmap_json,
                        strongest_correlations=strongest_correlations,
                        weakest_correlations=weakest_correlations,
                        launch_pad_stats=launch_pad_stats,
                        launch_pad=launch_pad_json,
                        performance_analysis=prediction_results,
                        bar_chart=bar_chart_json,
                        radar_chart=radar_chart_json,
                        radar_chart_social_sciences=radar_chart_social_sciences_json,
                        bar_chart_social_sciences=bar_chart_social_sciences_json
                        )  






#---------------------- Chat bot activation---------------------------------------------

@app.route('/chatbot')
def chatbot_page():
    return render_template('chatbot.html')

@app.route('/upload-image', methods=['POST'])
def upload_image():
    try:
        # User message
        message = request.form.get('message', '')

        # Get files
        if 'file' in request.files:
            file = request.files['file']
            if file:
                # Save to upload
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
                file.save(filepath)

                # Classificator
                extracted_text = ""
                file_extension = os.path.splitext(filepath)[1].lower()

                # Images (PNG, JPG, JPEG)
                if file_extension in ['.png', '.jpg', '.jpeg']:
                    try:
                        from PIL import Image
                        import pytesseract
                        img = Image.open(filepath)
                        extracted_text = pytesseract.image_to_string(img)
                    except Exception as e:
                        print(f"Error processing image: {e}")
                        extracted_text = "Error processing image."

                # TXT
                elif file_extension == '.txt':
                    try:
                        with open(filepath, 'r', encoding='utf-8') as txt_file:
                            extracted_text = txt_file.read()
                    except Exception as e:
                        print(f"Error processing text file: {e}")
                        extracted_text = "Error processing text file."

                # PDF
                elif file_extension == '.pdf':
                    try:
                        from PyPDF2 import PdfReader
                        reader = PdfReader(filepath)
                        for page in reader.pages:
                            extracted_text += page.extract_text()
                    except Exception as e:
                        print(f"Error processing PDF file: {e}")
                        extracted_text = "Error processing PDF file."

                # DOCX
                elif file_extension == '.docx':
                    try:
                        from docx import Document
                        doc = Document(filepath)
                        for paragraph in doc.paragraphs:
                            extracted_text += paragraph.text + '\n'
                    except Exception as e:
                        print(f"Error processing DOCX file: {e}")
                        extracted_text = "Error processing DOCX file."

                # XLSX
                elif file_extension == '.xlsx':
                    try:
                        import openpyxl
                        workbook = openpyxl.load_workbook(filepath)
                        for sheet_name in workbook.sheetnames:
                            sheet = workbook[sheet_name]
                            for row in sheet.iter_rows(values_only=True):
                                extracted_text += '\t'.join([str(cell) if cell else '' for cell in row]) + '\n'
                    except Exception as e:
                        print(f"Error processing XLSX file: {e}")
                        extracted_text = "Error processing XLSX file."

                else:
                    extracted_text = "Unsupported file type."

                # Create a context and send to ChatGPT
                full_context = f"User Message: {message}\nExtracted Text: {extracted_text}"
                print(f"Sending to OpenAI: {full_context}")

                # Send the request to GPT
                response = query_chatgpt(full_context)
                print(f"Response from OpenAI: {response}")
                return jsonify({'response': response}), 200

        # Just message ( not file attached )
        if message:
            print(f"Received message: {message}")
            response = query_chatgpt(message)
            return jsonify({'response': response}), 200

        # Files
        return jsonify({'error': 'No valid input provided'}), 400

    except Exception as e:
        print(f"Error in upload_image: {e}")
        return jsonify({'error': 'Internal Server Error', 'details': str(e)}), 500

# -------------------------------------------------------------------------------------------


if __name__ == "__main__":
    app.run(debug=True)
